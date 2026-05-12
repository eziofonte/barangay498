import os
import threading
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

DOCUMENTS = os.path.join(os.path.expanduser('~'), 'Documents')
REPORT_FOLDER = os.path.join(DOCUMENTS, 'Barangay498_Backups')


def run_monthly_report(app):
    from models import db, Senior, Transaction
    from fernet_crypto import decrypt

    try:
        with app.app_context():
            os.makedirs(REPORT_FOLDER, exist_ok=True)

            now = datetime.now()
            cutoff = now - timedelta(days=90)

            seniors = Senior.query.all()
            total_seniors = len(seniors)

            claimed_rows = db.session.query(Transaction.senior_id).filter(
                Transaction.date_released >= cutoff,
                Transaction.status == 'Released'
            ).distinct().all()
            claimed_ids = {sid for (sid,) in claimed_rows}
            claimed_count = len(claimed_ids)
            unclaimed_count = total_seniors - claimed_count

            period_txns = Transaction.query.filter(
                Transaction.date_released >= cutoff,
                Transaction.status == 'Released'
            ).all()
            direct_count = sum(1 for t in period_txns if t.release_type != 'Proxy')
            proxy_count = sum(1 for t in period_txns if t.release_type == 'Proxy')
            total_amount = sum(t.amount for t in period_txns)

            last_claim_by_senior = {}
            for t in Transaction.query.filter(Transaction.status == 'Released').all():
                cur = last_claim_by_senior.get(t.senior_id)
                if cur is None or t.date_released > cur:
                    last_claim_by_senior[t.senior_id] = t.date_released

            wb = Workbook()
            ws = wb.active
            ws.title = 'Summary'

            title_font = Font(bold=True, size=16, color='1A56A0')
            label_font = Font(bold=True)
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1A56A0', end_color='1A56A0', fill_type='solid')
            center = Alignment(horizontal='center')

            ws['A1'] = 'Barangay 498 — Monthly Summary'
            ws['A1'].font = title_font
            ws.merge_cells('A1:B1')
            ws['A2'] = f'Generated: {now.strftime("%B %d, %Y %I:%M %p")}'
            ws['A3'] = f'Period: Last 90 days (since {cutoff.strftime("%B %d, %Y")})'

            summary_rows = [
                ('Total Seniors Registered', total_seniors),
                ('Claimed This Period',      claimed_count),
                ('Unclaimed This Period',    unclaimed_count),
                ('Direct (Face) Releases',   direct_count),
                ('Proxy Releases',           proxy_count),
                ('Total Amount Released',    f'PHP {total_amount:,.2f}'),
            ]
            for i, (label, value) in enumerate(summary_rows, start=5):
                ws.cell(row=i, column=1, value=label).font = label_font
                ws.cell(row=i, column=2, value=value)

            ws.column_dimensions['A'].width = 32
            ws.column_dimensions['B'].width = 26

            ws2 = wb.create_sheet('Unclaimed Seniors')
            ws2.append(['Name', 'Age', 'Address', 'Last Claim Date'])
            for c in ws2[1]:
                c.font = header_font
                c.fill = header_fill
                c.alignment = center

            for s in seniors:
                if s.id in claimed_ids:
                    continue
                last = last_claim_by_senior.get(s.id)
                last_str = last.strftime('%b %d, %Y') if last else 'Never'
                try:
                    name = decrypt(s.full_name)
                    address = decrypt(s.address)
                except Exception:
                    name = s.full_name
                    address = s.address
                ws2.append([name, s.age, address, last_str])

            for col_letter, width in zip(['A', 'B', 'C', 'D'], [28, 6, 36, 16]):
                ws2.column_dimensions[col_letter].width = width

            timestamp = now.strftime('%Y-%m')
            filename = f'monthly_summary_{timestamp}.xlsx'
            path = os.path.join(REPORT_FOLDER, filename)
            wb.save(path)
            print(f'Monthly summary saved: {filename}')
    except Exception as e:
        print(f'Monthly report failed: {e}')


def schedule_monthly_report(app):
    def loop():
        last_run_month = None
        while True:
            now = datetime.now()
            if now.day == 1 and now.hour == 8 and last_run_month != (now.year, now.month):
                run_monthly_report(app)
                last_run_month = (now.year, now.month)
            threading.Event().wait(30 * 60)

    thread = threading.Thread(target=loop, daemon=True)
    thread.start()
    print('Monthly summary scheduler started (1st of month, 08:00)')
