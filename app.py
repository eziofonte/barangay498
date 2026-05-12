import json
import time as _time
import face_recognition
import numpy as np
import base64
import uuid as uuid_module
import os 
from backup import schedule_backups
from monthly_report import schedule_monthly_report
from fernet_crypto import encrypt, decrypt
from io import BytesIO
from PIL import Image
from flask import session
from flask import Flask, render_template, redirect, url_for, request, flash, Response
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from fileinput import filename
from models import db, Official, Senior, Transaction, ProxyEnrollment
from blink import detect_blink as check_blink
from datetime import datetime, timedelta


app = Flask(__name__)
app.config['STARTUP_TIME'] = str(_time.time())
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY', 'your-secret-key')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['UPLOAD_FOLDER'] = 'static/faces'
app.config['PERMANENT_SESSION_LIFETIME'] = 900
app.config['SESSION_PERMANENT'] = False

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

db.init_app(app)

schedule_backups()
schedule_monthly_report(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@login_manager.user_loader
def load_user(user_id):
    return Official.query.get(int(user_id))

@app.before_request
def check_session_timeout():
    if request.path.startswith('/static'):
        return
    if current_user.is_authenticated:
        if session.get('startup_time') != app.config.get('STARTUP_TIME'):
            logout_user()
            session.clear()
            return redirect(url_for('login'))
        last_active = session.get('last_active')
        if last_active:
            last_active_dt = datetime.fromisoformat(last_active)
            if datetime.now() - last_active_dt > timedelta(minutes=15):
                logout_user()
                session.clear()
                flash('Your session has expired due to inactivity. Please log in again.')
                return redirect(url_for('login'))
        session['last_active'] = datetime.now().isoformat()
        session.permanent = True

# --- Login ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    from datetime import datetime, timedelta

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        official = Official.query.filter_by(username=username).first()

        if official:
            # Check if account is locked
            if official.locked_until and datetime.now() < official.locked_until:
                remaining = int((official.locked_until - datetime.now()).total_seconds() / 60) + 1
                flash(f'Account locked. Too many failed attempts. Try again in {remaining} minute(s).')
                return render_template('login.html')

            if check_password_hash(official.password, password):
                # Reset failed attempts on success
                official.failed_attempts = 0
                official.locked_until = None
                db.session.commit()
                login_user(official, remember=False)
                session['startup_time'] = app.config.get('STARTUP_TIME')
                return redirect(url_for('index'))
            else:
                # Increment failed attempts
                official.failed_attempts += 1
                if official.failed_attempts >= 5:
                    official.locked_until = datetime.now() + timedelta(minutes=5)
                    official.failed_attempts = 0
                    db.session.commit()
                    flash('Too many failed attempts. Account locked for 5 minutes.')
                else:
                    db.session.commit()
                    remaining_attempts = 5 - official.failed_attempts
                    flash(f'Invalid password. {remaining_attempts} attempt(s) remaining.')
        else:
            flash('Invalid username or password.')

    return render_template('login.html')

# --- Logout ---
@app.route('/logout')
def logout():
    logout_user()
    session.clear()
    reason = request.args.get('reason')
    if reason == 'timeout':
        flash('You have been logged out due to inactivity.')
    return redirect(url_for('login'))

# --- Register Senior ---
@app.route('/register-senior', methods=['GET', 'POST'])
@login_required
def register_senior():
    if request.method == 'POST':
        import base64, uuid, time, io
        from PIL import Image as PILImage

        full_name = request.form.get('full_name')
        age = request.form.get('age')
        address = request.form.get('address')
        photo = request.files.get('photo')
        captured_photo = request.form.get('captured_photo')

        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

        photo_path = None
        img_array = None

        # Camera capture mode
        if captured_photo and captured_photo.startswith('data:image'):
            image_data = captured_photo.split(',')[1]
            image_bytes = base64.b64decode(image_data)
            filename = f"capture_{uuid.uuid4().hex}.jpg"
            photo_path = os.path.join(app.config['UPLOAD_FOLDER'], filename).replace('\\', '/')
            with open(photo_path, 'wb') as f:
                f.write(image_bytes)
            # Load from memory directly — avoids file locking
            pil_image = PILImage.open(io.BytesIO(image_bytes)).convert('RGB')
            img_array = np.array(pil_image)

        # Upload mode
        elif photo and allowed_file(photo.filename):
            filename = secure_filename(photo.filename)
            photo_path = os.path.join(app.config['UPLOAD_FOLDER'], filename).replace('\\', '/')
            photo.save(photo_path)
            time.sleep(0.3)
            pil_image = PILImage.open(photo_path).convert('RGB')
            img_array = np.array(pil_image)

        else:
            flash('Please provide a photo — either upload one or use the camera.')
            return render_template('register_senior.html')

        # Get encoding of new photo
        new_encodings = face_recognition.face_encodings(img_array)
        del img_array

        if not new_encodings:
            flash('No face detected in the photo. Please try again.')
            return render_template('register_senior.html')

        # Check if face already exists using cached encodings
        all_seniors = Senior.query.all()
        for existing_senior in all_seniors:
            try:
                if not existing_senior.face_encoding:
                    # Fallback: compute from file
                    pil = PILImage.open(existing_senior.photo_path).convert('RGB')
                    arr = np.array(pil)
                    encs = face_recognition.face_encodings(arr)
                    if not encs:
                        continue
                    known_encoding = encs[0]
                else:
                    known_encoding = np.array(json.loads(existing_senior.face_encoding))

                distance = face_recognition.face_distance([known_encoding], new_encodings[0])[0]
                results = face_recognition.compare_faces([known_encoding], new_encodings[0], tolerance=0.4)

                if results[0] and distance < 0.4:
                    flash(f'This person is already registered as {decrypt(existing_senior.full_name)}.')
                    return render_template('register_senior.html')
            except Exception as e:
                print(f"Encoding check error: {e}")
                continue

        # All checks passed — save senior
        senior = Senior(
            full_name=encrypt(full_name),
            age=age,
            address=encrypt(address),
            photo_path=photo_path,
            face_encoding=json.dumps(new_encodings[0].tolist())
        )
        db.session.add(senior)
        db.session.commit()
        flash('Senior registered successfully!')
        return redirect(url_for('index'))

    return render_template('register_senior.html')

# --- View All Seniors ---
@app.route('/seniors')
@login_required
def seniors():
    all_seniors = Senior.query.all()
    return render_template('seniors.html', seniors=all_seniors)

# --- Scan Page ---
@app.route('/scan')
@login_required
def scan():
    return render_template('scan.html')

# --- Generate Reference Number ---
def generate_reference():
    from datetime import datetime
    now = datetime.now()
    unique = uuid_module.uuid4().hex[:6].upper()
    return f"BRY-{now.strftime('%Y%m%d')}-{unique}"

def compute_and_save_encoding(senior):
    try:
        image = face_recognition.load_image_file(senior.photo_path)
        encodings = face_recognition.face_encodings(image)
        if encodings:
            import json
            senior.face_encoding = json.dumps(encodings[0].tolist())
            db.session.commit()
    except Exception:
        pass

# --- Face Recognition API ---
@app.route('/recognize', methods=['POST'])
@login_required
def recognize():
    data = request.get_json()
    image_data = data.get('image')

    if not image_data or not image_data.startswith('data:image'):
        return {'status': 'no_face', 'message': 'Invalid image data'}

    image_bytes = base64.b64decode(image_data.split(',')[1])
    new_image = face_recognition.load_image_file(BytesIO(image_bytes))
    new_encodings = face_recognition.face_encodings(new_image)

    if not new_encodings:
        return {'status': 'no_face', 'message': 'No face detected. Please try again.'}

    seniors = Senior.query.all()
    for senior in seniors:
        try:
            if not senior.face_encoding:
                continue
            known_encoding = np.array(json.loads(senior.face_encoding))
            distance = face_recognition.face_distance([known_encoding], new_encodings[0])[0]
            results = face_recognition.compare_faces([known_encoding], new_encodings[0], tolerance=0.4)

            if results[0] and distance < 0.4:
                cutoff = datetime.now() - timedelta(days=90)
                recent = Transaction.query.filter_by(senior_id=senior.id).filter(
                    Transaction.date_released >= cutoff,
                    Transaction.status == 'Released'
                ).first()

                if recent:
                    next_date = recent.date_released + timedelta(days=90)
                    return {
                        'status': 'already_claimed',
                        'message': f'Already claimed on {recent.date_released.strftime("%B %d, %Y")}. Next eligible: {next_date.strftime("%B %d, %Y")}'
                    }

                transaction = Transaction(
                    reference_number=generate_reference(),
                    senior_id=senior.id,
                    amount=1500.00,
                    released_by=current_user.name,
                    status='Pending'
                )
                db.session.add(transaction)
                db.session.commit()

                return {
                    'status': 'match',
                    'transaction_id': transaction.id,
                    'reference_number': transaction.reference_number,
                    'name': decrypt(senior.full_name),
                    'age': senior.age,
                    'address': decrypt(senior.address),
                    'photo': senior.photo_path.replace('\\', '/')
                }
        except Exception as e:
            print(f"Recognition error: {e}")
            continue

    return {'status': 'no_match', 'message': 'No matching senior found.'}

# --- Transaction History ---
@app.route('/history')
@login_required
def history():
    search = request.args.get('search', '')
    date_filter = request.args.get('date', '')

    query = Transaction.query.join(Senior)

    if search:
        query = query.filter(Senior.full_name.ilike(f'%{search}%'))

    if date_filter:
        from datetime import datetime
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d')
            query = query.filter(
                db.func.date(Transaction.date_released) == filter_date.date()
            )
        except:
            pass

    transactions = query.order_by(Transaction.date_released.desc()).all()
    return render_template('history.html', transactions=transactions, search=search, date_filter=date_filter)

# --- Export Analytics Report (Excel) ---
@app.route('/export-analytics', methods=['GET', 'POST'])
@login_required
def export_analytics():
    date_from_str = request.values.get('date_from')
    date_to_str   = request.values.get('date_to')

    if not date_from_str or not date_to_str:
        flash('Please provide both a start date and an end date.')
        return redirect(url_for('history'))

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
        date_to_inclusive = datetime.strptime(date_to_str, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        flash('Invalid date format. Please use the date pickers.')
        return redirect(url_for('history'))

    if date_from > date_to_inclusive:
        flash('"From" date must be before "To" date.')
        return redirect(url_for('history'))

    # Claimed: every Released transaction in range
    claimed_txs = Transaction.query.filter(
        Transaction.status == 'Released',
        Transaction.date_released >= date_from,
        Transaction.date_released <= date_to_inclusive
    ).order_by(Transaction.date_released.asc()).all()

    # Unclaimed: seniors with no Released transaction in range
    claimed_senior_ids = {t.senior_id for t in claimed_txs}
    if claimed_senior_ids:
        unclaimed_seniors = Senior.query.filter(~Senior.id.in_(claimed_senior_ids)).all()
    else:
        unclaimed_seniors = Senior.query.all()

    # ── Excel Workbook ─────────────────────────────────────────────────────
    wb = Workbook()

    title_font   = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    sub_font     = Font(name='Calibri', size=10, italic=True, color='4A5568')
    header_font  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    summary_font = Font(name='Calibri', size=11, bold=True, color='1A365D')

    claimed_title_fill    = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    claimed_header_fill   = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    unclaimed_title_fill  = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
    unclaimed_header_fill = PatternFill(start_color='B71C1C', end_color='B71C1C', fill_type='solid')
    alt_row_fill          = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    summary_fill          = PatternFill(start_color='EDF2F7', end_color='EDF2F7', fill_type='solid')

    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right',  vertical='center')

    thin   = Side(border_style='thin', color='CBD5E0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    range_label = f'Date Range: {date_from.strftime("%B %d, %Y")} to {datetime.strptime(date_to_str, "%Y-%m-%d").strftime("%B %d, %Y")}'
    generated_label = f'Generated: {datetime.now().strftime("%B %d, %Y %I:%M %p")}'
    subtitle_text = f'{range_label}  |  {generated_label}'

    def autofit(ws, headers, data_start_row=4):
        max_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(headers[col_idx - 1]))
            for r in range(data_start_row, max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is not None:
                    l = len(str(cell.value))
                    if l > max_len:
                        max_len = l
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    # ── Sheet 1: Claimed ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Claimed'

    headers1 = ['No.', 'Senior Name', 'Age', 'Address', 'Amount Released',
                'Date Released', 'Reference Number', 'Release Type',
                'Proxy Name', 'Released By']
    n_cols1 = len(headers1)

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols1)
    title_cell = ws1.cell(row=1, column=1, value='Claimed Allowance Report')
    title_cell.font = title_font
    title_cell.fill = claimed_title_fill
    title_cell.alignment = center
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols1)
    sub_cell = ws1.cell(row=2, column=1, value=subtitle_text)
    sub_cell.font = sub_font
    sub_cell.alignment = center
    ws1.row_dimensions[2].height = 18

    for col, h in enumerate(headers1, start=1):
        c = ws1.cell(row=4, column=col, value=h)
        c.font = header_font
        c.fill = claimed_header_fill
        c.alignment = center
        c.border = border
    ws1.row_dimensions[4].height = 24

    total_amount = 0.0
    for i, t in enumerate(claimed_txs, start=1):
        row_num = 4 + i
        senior = t.senior
        senior_name = senior.display_name if senior else ''
        senior_age = senior.age if senior else ''
        senior_addr = senior.display_address if senior else ''

        values = [
            i,
            senior_name,
            senior_age,
            senior_addr,
            float(t.amount or 0),
            t.date_released.strftime('%b %d, %Y %I:%M %p') if t.date_released else '',
            t.reference_number or '',
            t.release_type or 'Direct',
            t.proxy_name if (t.release_type == 'Proxy' and t.proxy_name) else '',
            t.released_by or ''
        ]
        for col, v in enumerate(values, start=1):
            c = ws1.cell(row=row_num, column=col, value=v)
            c.border = border
            if col in (2, 4, 9, 10):
                c.alignment = left_align
            else:
                c.alignment = center
            if col == 5:
                c.number_format = '"₱"#,##0.00'
            if i % 2 == 0:
                c.fill = alt_row_fill
        total_amount += float(t.amount or 0)

    summary_row1 = 4 + len(claimed_txs) + 1
    ws1.merge_cells(start_row=summary_row1, start_column=1, end_row=summary_row1, end_column=4)
    label1 = ws1.cell(row=summary_row1, column=1,
                      value=f'TOTAL  ({len(claimed_txs)} transaction{"s" if len(claimed_txs) != 1 else ""})')
    label1.font = summary_font
    label1.fill = summary_fill
    label1.alignment = right_align

    total_cell = ws1.cell(row=summary_row1, column=5, value=total_amount)
    total_cell.font = summary_font
    total_cell.fill = summary_fill
    total_cell.number_format = '"₱"#,##0.00'
    total_cell.alignment = center

    for col in range(6, n_cols1 + 1):
        c = ws1.cell(row=summary_row1, column=col, value='')
        c.fill = summary_fill
    for col in range(1, n_cols1 + 1):
        ws1.cell(row=summary_row1, column=col).border = border
    ws1.row_dimensions[summary_row1].height = 22

    autofit(ws1, headers1)

    # ── Sheet 2: Unclaimed ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('Unclaimed')

    headers2 = ['No.', 'Senior Name', 'Age', 'Address', 'Status']
    n_cols2 = len(headers2)

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols2)
    title2 = ws2.cell(row=1, column=1, value='Unclaimed Allowance Report')
    title2.font = title_font
    title2.fill = unclaimed_title_fill
    title2.alignment = center
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols2)
    sub2 = ws2.cell(row=2, column=1, value=subtitle_text)
    sub2.font = sub_font
    sub2.alignment = center
    ws2.row_dimensions[2].height = 18

    for col, h in enumerate(headers2, start=1):
        c = ws2.cell(row=4, column=col, value=h)
        c.font = header_font
        c.fill = unclaimed_header_fill
        c.alignment = center
        c.border = border
    ws2.row_dimensions[4].height = 24

    for i, sr in enumerate(unclaimed_seniors, start=1):
        row_num = 4 + i
        values = [
            i,
            sr.display_name,
            sr.age,
            sr.display_address,
            'Not Yet Claimed'
        ]
        for col, v in enumerate(values, start=1):
            c = ws2.cell(row=row_num, column=col, value=v)
            c.border = border
            c.alignment = left_align if col in (2, 4) else center
            if i % 2 == 0:
                c.fill = alt_row_fill

    summary_row2 = 4 + len(unclaimed_seniors) + 1
    ws2.merge_cells(start_row=summary_row2, start_column=1, end_row=summary_row2, end_column=4)
    label2 = ws2.cell(row=summary_row2, column=1,
                      value=f'TOTAL UNCLAIMED  ({len(unclaimed_seniors)} senior{"s" if len(unclaimed_seniors) != 1 else ""})')
    label2.font = summary_font
    label2.fill = summary_fill
    label2.alignment = right_align

    count_cell = ws2.cell(row=summary_row2, column=5, value=len(unclaimed_seniors))
    count_cell.font = summary_font
    count_cell.fill = summary_fill
    count_cell.alignment = center

    for col in range(1, n_cols2 + 1):
        ws2.cell(row=summary_row2, column=col).border = border
    ws2.row_dimensions[summary_row2].height = 22

    autofit(ws2, headers2)

    # ── Send file ──────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'analytics_{date_from.strftime("%Y%m%d")}_to_{datetime.strptime(date_to_str, "%Y-%m-%d").strftime("%Y%m%d")}.xlsx'
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

# --- Edit Senior ---
@app.route('/seniors/<int:senior_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_senior(senior_id):
    senior = Senior.query.get_or_404(senior_id)
    if request.method == 'POST':
        senior.full_name = encrypt(request.form.get('full_name'))
        senior.age = request.form.get('age')
        senior.address = encrypt(request.form.get('address'))

        photo = request.files.get('photo')
        if photo and allowed_file(photo.filename):
            filename = secure_filename(photo.filename)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            photo_path = os.path.join(app.config['UPLOAD_FOLDER'], filename).replace('\\', '/')
            photo.save(photo_path)
            senior.photo_path = photo_path

        db.session.commit()
        if photo and allowed_file(photo.filename):
            compute_and_save_encoding(senior)
        flash('Senior updated successfully!')
        return redirect(url_for('seniors'))

    return render_template('edit_senior.html', senior=senior)

# --- Delete Senior ---
@app.route('/seniors/<int:senior_id>/delete', methods=['POST'])
@login_required
def delete_senior(senior_id):
    senior = Senior.query.get_or_404(senior_id)
    Transaction.query.filter_by(senior_id=senior.id).delete()
    db.session.delete(senior)
    db.session.commit()
    flash('Senior removed from the system.', 'info')
    return redirect(url_for('seniors'))

# --- Reset Senior Claim ---
@app.route('/transactions/<int:transaction_id>/reset', methods=['POST'])
@login_required
def reset_claim(transaction_id):
    transaction = Transaction.query.get_or_404(transaction_id)
    db.session.delete(transaction)
    db.session.commit()
    flash('Claim reset successfully. Senior can now claim again.')
    return redirect(url_for('history'))

# --- Reset All Claims ---
@app.route('/transactions/reset-all', methods=['POST'])
@login_required
def reset_all_claims():
    from datetime import datetime, timedelta
    cutoff = datetime.now() - timedelta(days=90)
    transactions = Transaction.query.filter(
        Transaction.date_released >= cutoff,
        Transaction.status == 'Released'
    ).all()
    for t in transactions:
        t.status = 'Reset'
    db.session.commit()
    flash('All claims have been reset. All seniors can now claim again.')
    return redirect(url_for('seniors'))

# --- DSS Page ---
@app.route('/')
@login_required
def index():
    from datetime import datetime, timedelta
    now = datetime.now()
    cutoff = now - timedelta(days=90)

    senior_count = Senior.query.count()
    transaction_count = Transaction.query.count()
    total_released = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.status == 'Released'
    ).scalar() or 0

    claimed_ids = db.session.query(Transaction.senior_id).filter(
        Transaction.date_released >= cutoff,
        Transaction.status == 'Released'
    ).distinct().all()
    claimed_ids = [c[0] for c in claimed_ids]
    claimed_count = len(claimed_ids)
    unclaimed_count = senior_count - claimed_count

    unclaimed_seniors = Senior.query.filter(
        ~Senior.id.in_(claimed_ids)
    ).all() if claimed_ids else Senior.query.all()

    last_claim_by_senior = {}
    for t in Transaction.query.filter(Transaction.status == 'Released').all():
        if t.date_released is None:
            continue
        cur = last_claim_by_senior.get(t.senior_id)
        if cur is None or t.date_released > cur:
            last_claim_by_senior[t.senior_id] = t.date_released

    priority_list = []
    for s in unclaimed_seniors:
        last = last_claim_by_senior.get(s.id)
        if last is None:
            days_since = 999
            last_str = 'Never'
        else:
            days_since = max(0, (now - last).days)
            last_str = last.strftime('%b %d, %Y')
        age = s.age or 0
        score = round((days_since * 0.7) + (age * 0.3), 1)
        priority_list.append({
            'senior': s,
            'last_claim_str': last_str,
            'days_since': days_since,
            'score': score,
        })
    priority_list.sort(key=lambda x: x['score'], reverse=True)

    insights = []
    if unclaimed_count > 0:
        insights.append(f'{unclaimed_count} senior(s) have not claimed their allowance in the last 90 days.')
    if unclaimed_count == 0 and senior_count > 0:
        insights.append('All registered seniors have claimed their allowance. Great job!')
    if senior_count == 0:
        insights.append('No seniors are registered yet. Start by registering senior citizens.')
    if claimed_count > 0 and unclaimed_count > 0:
        rate = (claimed_count / senior_count) * 100
        insights.append(f'Current claim rate is {rate:.1f}%. Consider reaching out to unclaimed seniors.')

    return render_template('index.html',
        senior_count=senior_count,
        transaction_count=transaction_count,
        total_released=total_released,
        claimed_count=claimed_count,
        unclaimed_count=unclaimed_count,
        unclaimed_seniors=unclaimed_seniors,
        priority_list=priority_list,
        insights=insights,
        now=now
    )

@app.route('/confirm-release', methods=['POST'])
@login_required
def confirm_release():
    import base64, uuid as uuid_lib
    data = request.get_json()
    transaction_id = data.get('transaction_id')
    signature_data = data.get('signature')
    release_photo_data = data.get('release_photo')

    transaction = Transaction.query.get_or_404(transaction_id)

    os.makedirs('static/signatures', exist_ok=True)
    os.makedirs('static/release_photos', exist_ok=True)

    # Save signature
    if signature_data and signature_data.startswith('data:image'):
        sig_bytes = base64.b64decode(signature_data.split(',')[1])
        senior_name = transaction.senior.full_name.replace(' ', '_')
        ref = transaction.reference_number
        sig_filename = f"{senior_name}_{ref}_signature.png"
        sig_path = 'static/signatures/' + sig_filename
        with open(sig_path, 'wb') as f:
            f.write(sig_bytes)
        transaction.signature_path = sig_path

    # Save release photo
    if release_photo_data and release_photo_data.startswith('data:image'):
        photo_bytes = base64.b64decode(release_photo_data.split(',')[1])
        photo_filename = f"{senior_name}_{ref}_release.jpg"
        photo_path = 'static/release_photos/' + photo_filename
        with open(photo_path, 'wb') as f:
            f.write(photo_bytes)
        transaction.release_photo_path = photo_path

    transaction.status = 'Released'
    db.session.commit()

    return {'status': 'success', 'reference_number': transaction.reference_number}

# --- Get Seniors List for Proxy ---
@app.route('/seniors-list')
@login_required
def seniors_list():
    seniors = Senior.query.all()
    return {'seniors': [{'id': s.id, 'name': decrypt(s.full_name), 'age': s.age, 'address': decrypt(s.address), 'photo': s.photo_path.replace('\\', '/')} for s in seniors]}

# --- Reset Individual Senior Claim ---
@app.route('/seniors/<int:senior_id>/reset-claim', methods=['POST'])
@login_required
def reset_senior_claim(senior_id):
    from datetime import datetime, timedelta
    cutoff = datetime.now() - timedelta(days=90)
    Transaction.query.filter_by(senior_id=senior_id).filter(
        Transaction.date_released >= cutoff,
        Transaction.status == 'Released'
    ).delete()
    db.session.commit()
    senior = Senior.query.get_or_404(senior_id)
    flash(f'Claim reset for {senior.display_name}. They can now claim again.')
    return redirect(url_for('seniors'))

# --- Blink Detection ---
@app.route('/detect-blink', methods=['POST'])
@login_required
def detect_blink_route():
    data = request.get_json()
    image_data = data.get('image')
    reset = data.get('reset', False)

    if reset:
        from blink import reset_blink_counter
        reset_blink_counter()
        return {'blink': False, 'face': False}

    if not image_data or not image_data.startswith('data:image'):
        return {'blink': False, 'face': False}

    image_bytes = base64.b64decode(image_data.split(',')[1])
    result = check_blink(image_bytes)
    return result

# --- Proxy Release ---
@app.route('/proxy-release', methods=['POST'])
@login_required
def proxy_release():
    from werkzeug.security import check_password_hash
    data = request.get_json()
    captain_pin         = data.get('captain_pin')
    senior_id           = data.get('senior_id')
    proxy_name          = data.get('proxy_name')
    proxy_relationship  = data.get('proxy_relationship')
    proxy_enrollment_id = data.get('proxy_enrollment_id')
    signature_data      = data.get('signature')
    release_photo_data  = data.get('release_photo')

    # Verify captain PIN
    captain = Official.query.filter_by(role='captain').first()
    if not captain or not captain.captain_pin:
        return {'status': 'error', 'message': 'No captain account found.'}
    if not check_password_hash(captain.captain_pin, captain_pin):
        return {'status': 'error', 'message': 'Incorrect captain PIN.'}

    senior = Senior.query.get_or_404(senior_id)

    # 90-day cooldown check
    cutoff = datetime.now() - timedelta(days=90)
    recent = Transaction.query.filter_by(senior_id=senior.id).filter(
        Transaction.date_released >= cutoff,
        Transaction.status == 'Released'
    ).first()
    if recent:
        next_date = recent.date_released + timedelta(days=90)
        return {
            'status': 'already_claimed',
            'message': f'Already claimed on {recent.date_released.strftime("%B %d, %Y")}. Next eligible: {next_date.strftime("%B %d, %Y")}'
        }

    transaction = Transaction(
        reference_number=generate_reference(),
        senior_id=senior.id,
        amount=1500.00,
        released_by=current_user.name,
        status='Released',
        release_type='Proxy',
        proxy_name=proxy_name,
        proxy_relationship=proxy_relationship,
        proxy_enrollment_id=proxy_enrollment_id if proxy_enrollment_id else None
    )
    db.session.add(transaction)
    db.session.flush()

    os.makedirs('static/signatures', exist_ok=True)
    os.makedirs('static/release_photos', exist_ok=True)

    senior_name = senior.full_name.replace(' ', '_')
    ref = transaction.reference_number

    if signature_data and signature_data.startswith('data:image'):
        sig_bytes = base64.b64decode(signature_data.split(',')[1])
        sig_filename = f"{senior_name}_{ref}_proxy_signature.png"
        sig_path = 'static/signatures/' + sig_filename
        with open(sig_path, 'wb') as f:
            f.write(sig_bytes)
        transaction.signature_path = sig_path

    if release_photo_data and release_photo_data.startswith('data:image'):
        photo_bytes = base64.b64decode(release_photo_data.split(',')[1])
        photo_filename = f"{senior_name}_{ref}_proxy_release.jpg"
        photo_path = 'static/release_photos/' + photo_filename
        with open(photo_path, 'wb') as f:
            f.write(photo_bytes)
        transaction.release_photo_path = photo_path

    db.session.commit()

    return {
        'status': 'success',
        'reference_number': transaction.reference_number,
        'senior_name': senior.full_name,
        'proxy_name': proxy_name
    }

@app.route('/verify-captain-pin', methods=['POST'])
@login_required
def verify_captain_pin():
    from werkzeug.security import check_password_hash
    data = request.get_json()
    captain_pin = data.get('captain_pin')
    captain = Official.query.filter_by(role='captain').first()
    if not captain or not captain.captain_pin:
        return {'status': 'error', 'message': 'No captain account found.'}
    if not check_password_hash(captain.captain_pin, captain_pin):
        return {'status': 'error', 'message': 'Incorrect captain PIN.'}
    return {'status': 'ok'}

@app.route('/fix-captain')
@login_required
def fix_captain():
    admin = Official.query.filter_by(username='admin').first()
    admin.role = 'captain'
    admin.captain_pin = generate_password_hash('captain1234')
    db.session.commit()
    return {'status': 'done', 'message': 'Admin is now captain with PIN: captain1234'}

@app.route('/admin/reset-captain-pin', methods=['GET', 'POST'])
@login_required
def reset_captain_pin():
    message = None
    error = None

    if request.method == 'POST':
        admin_password = request.form.get('admin_password')
        new_pin = (request.form.get('new_pin') or '').strip()
        confirm_pin = request.form.get('confirm_pin')

        if not check_password_hash(current_user.password, admin_password):
            error = 'Incorrect admin password.'
        elif len(new_pin) < 4:
            error = 'PIN must be at least 4 characters.'
        elif new_pin != confirm_pin:
            error = 'PINs do not match.'
        else:
            captain = Official.query.filter_by(role='captain').first()
            if not captain:
                error = 'No captain account found.'
            else:
                captain.captain_pin = generate_password_hash(new_pin)
                db.session.commit()
                message = 'Captain PIN successfully updated.'

    return render_template('reset_captain_pin.html', message=message, error=error)

# ── Proxy Management ──────────────────────────────────────────────────────────

@app.route('/proxy-management')
@login_required
def proxy_management():
    if current_user.role not in ('captain', 'admin'):
        return redirect(url_for('index'))
    seniors     = Senior.query.order_by(Senior.full_name).all()
    enrollments = ProxyEnrollment.query.order_by(ProxyEnrollment.enrolled_at.desc()).all()
    return render_template('proxy_management.html', seniors=seniors, enrollments=enrollments)


@app.route('/enroll-proxy', methods=['POST'])
@login_required
def enroll_proxy():
    if current_user.role not in ('captain', 'admin'):
        return redirect(url_for('index'))

    senior_id    = request.form.get('senior_id')
    full_name    = (request.form.get('full_name')    or '').strip()
    relationship = (request.form.get('relationship') or '').strip()
    id_type      = (request.form.get('id_type')      or '').strip()
    id_number    = (request.form.get('id_number')    or '').strip()

    if not all([senior_id, full_name, relationship, id_type, id_number]):
        flash('All fields are required.', 'error')
        return redirect(url_for('proxy_management'))

    os.makedirs('static/proxy_ids',    exist_ok=True)
    os.makedirs('static/proxy_faces',  exist_ok=True)

    def save_upload(field, folder, prefix):
        f = request.files.get(field)
        if not f or not f.filename:
            return None
        ext      = f.filename.rsplit('.', 1)[-1].lower()
        filename = f"{prefix}_{uuid_module.uuid4().hex[:8]}.{ext}"
        path     = os.path.join(folder, filename)
        f.save(path)
        return path

    id_photo_path   = save_upload('id_photo',   'static/proxy_ids',   f"id_{senior_id}")
    face_photo_path = save_upload('face_photo', 'static/proxy_faces', f"face_{senior_id}")

    if not id_photo_path:
        flash('ID photo is required.', 'error')
        return redirect(url_for('proxy_management'))

    enrollment = ProxyEnrollment(
        senior_id    = senior_id,
        full_name=encrypt(full_name),
        relationship = relationship,
        id_type      = id_type,
        id_number=encrypt(id_number),
        id_photo     = id_photo_path,
        face_photo   = face_photo_path,
        enrolled_by  = current_user.id
    )
    db.session.add(enrollment)
    db.session.commit()
    flash(f'{full_name} enrolled as proxy successfully.', 'success')
    return redirect(url_for('proxy_management'))


@app.route('/proxy-enrollment/<int:enrollment_id>/toggle', methods=['POST'])
@login_required
def toggle_proxy_enrollment(enrollment_id):
    if current_user.role not in ('captain', 'admin'):
        return redirect(url_for('index'))
    enrollment           = ProxyEnrollment.query.get_or_404(enrollment_id)
    enrollment.is_active = not enrollment.is_active
    db.session.commit()
    return redirect(url_for('proxy_management'))


@app.route('/proxy-enrollment/<int:enrollment_id>/delete', methods=['POST'])
@login_required
def delete_proxy_enrollment(enrollment_id):
    if current_user.role not in ('captain', 'admin'):
        return redirect(url_for('index'))
    enrollment = ProxyEnrollment.query.get_or_404(enrollment_id)
    db.session.delete(enrollment)
    db.session.commit()
    flash('Proxy enrollment removed.', 'success')
    return redirect(url_for('proxy_management'))


@app.route('/enrolled-proxies/<int:senior_id>')
@login_required
def enrolled_proxies(senior_id):
    proxies = ProxyEnrollment.query.filter_by(senior_id=senior_id, is_active=True).all()
    return {
        'proxies': [{
            'id':           p.id,
            'full_name':    decrypt(p.full_name),
            'relationship': p.relationship,
            'face_photo':   p.face_photo.replace('\\', '/') if p.face_photo else None
        } for p in proxies]
    }


if __name__ == '__main__':
    app.run(debug=True)